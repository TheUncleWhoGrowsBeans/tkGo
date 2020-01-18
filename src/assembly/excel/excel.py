#!/usr/bin/env python
# coding=utf-8

# @Author              : Uncle Bean
# @Date                : 2020-01-16 09:38:42
# @LastEditors: Uncle Bean
# @LastEditTime: 2020-01-18 20:22:47
# @FilePath            : \src\utils\file\excel.py
# @Description         : 

import os
import requests
import openpyxl
import validators
from threading import Thread
from traceback import print_exc


class Excel(object):

    IMG_DOWNLOAD_EXISTS = "图片已存在"
    IMG_DOWNLOAD_NOT_EXISTS = "图片不存在"
    IMG_DOWNLOAD_DONE = "图片已下载"
    IMG_DOWNLOAD_ERROR = "图片下载失败"
    IMG_URL_ERROR = "不是有效图片链接"
    IMG_HEIGHT = 120
    IMG_WIDTH = 120
    IMG_CELL_HEIGHT = 90
    IMG_CELL_WIDTH = 15

    def __init__(self, path, img_dir="", img_dispersed=True, img_ext="png,jpg,jpeg", stdout=print):
        """Excel小工具

        :param path: Excel路径
        :param img_dir: 图片目录，用户保存Excel中的图片
        :param img_dispersed: True则根据url分目录存储图片，False则存储在单个文件夹中
        :param img_ext: 图片文件后缀，用户识别Excel中的图片链接
        :param stdout: 信息输出函数，用于信息输出的重定向
        """
        self.stdout = stdout
        self.img_dir = img_dir
        self.img_dispersed = img_dispersed
        self.img_ext = list(map(lambda  x: x if x.startswith(".") else "." + x, img_ext.split(",")))  # 图片文件后缀名
        self.img_download_failed = dict()
        self.img_add_failed = dict()
        self.path = path
        
        self.wb = openpyxl.load_workbook(self.path)
        self.sheet_names = self.wb.sheetnames

        self.stdout("img_ext ->", self.img_ext)
        self.stdout("sheet_names ->", self.wb.sheetnames)

    def is_img_url(self, value: str):
        if value.startswith("=HYPERLINK"): value = value[12:-2]  # 去掉excel超链接函数
        is_img_url = False
        if validators.url(value):  # 判断是否是url
            for img_ext in self.img_ext:
                if value.lower().endswith(img_ext):
                    is_img_url = True
                    break
        return is_img_url, value
    
    def check_and_mkdir(self, path=None, dir=None, max_num_cycles=100):
        """检查目录是否存在，若不存在则创建
        """
        dir_name_list = list()
        dir_name = dir if dir else os.path.dirname(path)
        cycle_times = 0
        while not os.path.exists(dir_name):
            dir_name_list.append(dir_name)
            dir_name = os.path.dirname(dir_name)
            cycle_times += 1
            if cycle_times >= max_num_cycles:
                raise RuntimeError('CheckDirError -> {}'.format(path))
        while len(dir_name_list) > 0:
            dir_name_to_make = dir_name_list.pop()
            os.makedirs(dir_name_to_make)

    def img_url_to_path(self, img_url):
        img_name = img_url.replace(r"https://", "").replace(r"http://", "").replace(r":", "#")
        if not self.img_dispersed: img_name = img_name.replace("/", "-")
        return os.path.abspath(os.path.join(self.img_dir,img_name))

    def download_img(self, img_url, sheet_name=None, coordinate=None, timeout=15):
        img_path = self.img_url_to_path(img_url)
        if os.path.exists(img_path):
            self.stdout(sheet_name, coordinate, img_url, self.IMG_DOWNLOAD_EXISTS)
        else:
            self.check_and_mkdir(path=img_path)
            r = requests.get(img_url, timeout=timeout)
            if not r.content[:4] == b'\xff\xd8\xff\xe0': 
                self.stdout(sheet_name, coordinate, img_url, self.IMG_DOWNLOAD_ERROR, self.IMG_URL_ERROR)
                return False
            with open(img_path, "wb") as f:
                f.write(r.content)
            self.stdout(sheet_name, coordinate, img_url, self.IMG_DOWNLOAD_DONE)
        return True
                                    
    def download_img_of_sheet(self, sheet_name, try_num=3):
        sheet = self.wb[sheet_name]
        row_num = sheet.max_row
        column_num = sheet.max_column
        self.stdout(sheet_name, "row_num ->", row_num, " column_num ->", column_num)

        for row in sheet.rows:
            for cell in row:
                cell_value = str(cell.value)
                is_img_url, img_url = self.is_img_url(cell_value)
                if is_img_url:
                    try_id = 0
                    while try_id < try_num:
                        try:
                            try_id += 1
                            if not self.download_img(img_url, sheet_name, cell.coordinate): 
                                self.img_download_failed[img_url] = (sheet_name + " " + cell.coordinate, self.IMG_URL_ERROR)
                            break
                        except Exception as e:
                            if try_id == try_num:
                                self.img_download_failed[img_url] = (sheet_name + " " + cell.coordinate, str(e))
                                self.stdout(img_url, self.IMG_DOWNLOAD_ERROR, str(e))

    def download_img_of_wb(self):
        threads = list()
        for sheet_name in self.sheet_names:
            threads.append(Thread(target=self.download_img_of_sheet, args=(sheet_name,)))
        for t in threads:
            t.start()
        for t in threads:
            t.join()
    
    def add_img_of_sheet(self, sheet_name):
        sheet = self.wb[sheet_name]
        for row in sheet.rows:
            for cell in row:
                cell_value = str(cell.value)
                is_img_url, img_url = self.is_img_url(cell_value)
                if is_img_url:
                    img_path = self.img_url_to_path(img_url)
                    if os.path.exists(img_path):
                        try:
                            img = openpyxl.drawing.image.Image(img_path)
                            img.height = self.IMG_HEIGHT
                            img.width = self.IMG_WIDTH
                            sheet.add_image(img, cell.coordinate)
                            sheet.row_dimensions[cell.row].height = self.IMG_CELL_HEIGHT
                            column_letter = openpyxl.utils.get_column_letter(cell.column)
                            sheet.column_dimensions[column_letter].width = self.IMG_CELL_WIDTH
                            cell.alignment =  openpyxl.styles.Alignment(wrapText=True)
                        except Exception as e:
                            print_exc()
                            self.img_add_failed[img_url] = (sheet_name + " " + cell.coordinate, str(e))
                            self.stdout(sheet_name, cell.coordinate, img_url, str(e))
                    else:
                        self.img_add_failed[img_url] = (sheet_name + " " + cell.coordinate, self.IMG_DOWNLOAD_NOT_EXISTS)
                        self.stdout(sheet_name, cell.coordinate, img_url, self.IMG_DOWNLOAD_NOT_EXISTS)
                        
    def save(self, path):
        self.wb.save(filename=path)
        
    def add_img_of_wb(self):
        threads = list()
        for sheet_name in self.sheet_names:
            threads.append(Thread(target=self.add_img_of_sheet, args=(sheet_name,)))
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        
        path_split = os.path.splitext(self.path)
        path_excel_with_img = path_split[0] + ".image" + path_split[1]
        self.save(path=path_excel_with_img)


if __name__ == "__main__":
    excel = Excel(r"C:\easonhu\proj\jollychic\atlas\AdHoc\2020\test\test.xlsx", img_dispersed=True)

    # excel.download_img_of_sheet(excel.sheet_names[0])

    excel.download_img_of_wb()
    excel.add_img_of_wb()
